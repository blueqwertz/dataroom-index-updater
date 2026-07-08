#!/usr/bin/env python3
"""
Dataroom-Index-Updater
======================

Fuegt neue Dokumente aus einem frischen Datenraum-Export ("input") in einen
bestehenden, angereicherten Index ("current") ein.

Idee:
- Der current-Index ist normalerweise der Datenraum-Index + eigene Zusatzspalten
  (z.B. "Allocation", "Comment"). Spalten werden ueber die Kopfzeile (Header)
  zwischen Input und Current abgeglichen -- Zusatzspalten bleiben erhalten.
- Neue Zeilen (Dokumente/Ordner), die im Input vorkommen, im Current aber fehlen,
  werden eingefuegt, mit den Input-Daten befuellt und gelb markiert.
- Bestehende Zeilen im Current bleiben UNVERAENDERT (inkl. alter Index-Nummern,
  Allocation, Comments, Formatierung). Es wird nur hinzugefuegt, nie ueberschrieben
  oder geloescht.

Der "Anker":
- Um zu erkennen welche Zeilen neu sind UND an welche Stelle sie gehoeren, braucht
  es einen stabilen Schluessel ("Anker"), der eine Zeile ueber Versionen hinweg
  identifiziert. Das Skript sucht diesen automatisch:
    1. Bevorzugt eine echte ID-Spalte (Werte vollstaendig + eindeutig in beiden
       Dateien; Header-Hinweise wie "id"/"unique" geben Bonus).
    2. Faellt zurueck auf einen zusammengesetzten Schluessel aus STABILEN Textspalten
       (Title, Type, File Type, Fileroom, Level, Hyperlink) -- bewusst OHNE Index,
       Dateigroesse und Datum, da sich diese zwischen Versionen aendern koennen.
    3. Notfalls "Title" allein, mit Warnung.
- Eine neue Input-Zeile wird direkt hinter ihren "Anker" gesetzt: die naechste
  darueberliegende Input-Zeile, die auch im Current existiert. So landet das Dok
  an der richtigen Stelle, auch wenn sich die Index-Nummerierung verschoben hat.

Aufruf:
    python update_index.py                 # nutzt Standard-Dateinamen
    python update_index.py --input in.xlsx --current cur.xlsx --output out.xlsx
"""

import argparse
import sys
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")

# Spalten, die sich zwischen Versionen aendern koennen -> nie als Schluessel nutzen
UNSTABLE_HEADERS = {"index", "file size in mb", "date available", "status"}
# Bevorzugte, stabile Spalten fuer einen zusammengesetzten Schluessel (Reihenfolge = Prioritaet)
STABLE_PREFERENCE = ["unique id", "title", "hyperlink", "type", "file type", "fileroom", "level"]


def norm(h):
    return str(h).strip().lower() if h is not None else ""


def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def read_sheet(path):
    """Liest ein Blatt: gibt (workbook, worksheet, header_map, rows) zurueck.
    header_map: normalisierter Header -> Spaltenindex (1-basiert)
    rows: Liste von row-Indizes (1-basiert) der Datenzeilen (ohne Kopfzeile)
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_map = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if h:
            header_map[h] = c
    rows = list(range(2, ws.max_row + 1))
    # leere Zeilen am Ende ignorieren
    rows = [r for r in rows if any(not is_empty(ws.cell(r, c).value)
                                   for c in range(1, ws.max_column + 1))]
    return wb, ws, header_map, rows


def col_stats(ws, col, rows):
    vals = [ws.cell(r, col).value for r in rows]
    nonempty = [v for v in vals if not is_empty(v)]
    complete = len(nonempty) == len(vals) and len(vals) > 0
    unique = len(set(map(str, nonempty))) == len(nonempty) and len(nonempty) > 0
    return complete, unique


def choose_key(in_map, in_ws, in_rows, cur_map, cur_ws, cur_rows):
    """Waehlt die Anker-Spalten. Gibt eine Liste normalisierter Header zurueck,
    die in beiden Dateien existieren und zusammen einen eindeutigen Schluessel bilden.
    """
    shared = [h for h in in_map if h in cur_map]

    # 1) Beste Einzelspalte (echte ID)
    best = None  # (score, header)
    for h in shared:
        ic, iu = col_stats(in_ws, in_map[h], in_rows)
        cc, cu = col_stats(cur_ws, cur_map[h], cur_rows)
        if ic and iu and cc and cu:
            score = 100
            if "id" in h:
                score += 50
            if "unique" in h:
                score += 50
            if h in UNSTABLE_HEADERS:
                score -= 200  # eindeutig aber instabil -> nur Notnagel
            if best is None or score > best[0]:
                best = (score, h)
    if best and best[0] > 0:
        return [best[1]], "einzelne ID-Spalte"

    # 2) Zusammengesetzter Schluessel aus stabilen Spalten
    ordered = [h for h in STABLE_PREFERENCE if h in shared and h not in UNSTABLE_HEADERS]
    ordered += [h for h in shared if h not in ordered and h not in UNSTABLE_HEADERS]

    key = []
    for h in ordered:
        key.append(h)
        if key_is_unique(in_ws, in_map, in_rows, key) and \
           key_is_unique(cur_ws, cur_map, cur_rows, key):
            return key, "zusammengesetzter Schluessel"

    # 3) Notnagel
    if "title" in shared:
        return ["title"], "NOTFALL: nur 'Title' (nicht garantiert eindeutig)"
    if shared:
        return [shared[0]], f"NOTFALL: '{shared[0]}' (nicht garantiert eindeutig)"
    raise SystemExit("Keine gemeinsamen Spalten zwischen Input und Current gefunden.")


def key_is_unique(ws, hmap, rows, headers):
    seen = set()
    for r in rows:
        k = row_key(ws, hmap, r, headers)
        if k in seen:
            return False
        seen.add(k)
    return True


def row_key(ws, hmap, row, headers):
    return tuple(norm(ws.cell(row, hmap[h]).value) for h in headers)


def build_col_templates(ws, header_map, rows):
    """Pro Spalte eine Vorlagen-Formatierung (StyleArray) bestimmen.
    Bevorzugt eine nicht-leere DOCUMENT-Zeile (korrekte Zahlen-/Datumsformate).
    """
    type_col = header_map.get("type")
    templates = {}
    for c in range(1, ws.max_column + 1):
        chosen = None
        fallback = None
        for r in rows:
            if is_empty(ws.cell(r, c).value):
                continue
            if fallback is None:
                fallback = ws.cell(r, c)
            if type_col and norm(ws.cell(r, type_col).value) == "document":
                chosen = ws.cell(r, c)
                break
        src = chosen or fallback or ws.cell(1, c)
        templates[c] = copy(src._style)
    return templates


def capture_row(ws, row):
    """Werte + Styles einer Zeile festhalten (fuer verlustfreies Neuschreiben)."""
    return {c: (ws.cell(row, c).value, copy(ws.cell(row, c)._style))
            for c in range(1, ws.max_column + 1)}


def parse_highlight_cols(spec, max_col):
    """Parst eine Spaltenangabe wie "B-D", "B:D", "A,C-E" oder "2-4" in ein
    Set von 1-basierten Spaltenindizes. Leer/None/"all" -> alle Spalten."""
    if spec is None or str(spec).strip() == "" or str(spec).strip().lower() in ("all", "alle"):
        return set(range(1, max_col + 1))
    cols = set()
    for part in str(spec).replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        sep = "-" if "-" in part else (":" if ":" in part else None)
        if sep:
            a, b = (p.strip() for p in part.split(sep, 1))
            lo, hi = _col_index(a), _col_index(b)
            if lo > hi:
                lo, hi = hi, lo
            cols.update(range(lo, hi + 1))
        else:
            cols.add(_col_index(part))
    return {c for c in cols if 1 <= c <= max_col}


def _col_index(token):
    token = token.strip()
    if token.isdigit():
        return int(token)
    return column_index_from_string(token.upper())


def shared_headers(path_a, path_b):
    """Gemeinsame Header beider Dateien (Original-Schreibweise aus Datei A),
    fuer die Anker-Auswahl in der GUI."""
    def headers(path):
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        wb.close()
        return {norm(v): str(v).strip() for v in row if not is_empty(v)}
    a, b = headers(path_a), headers(path_b)
    return [a[h] for h in a if h in b]


def run_update(input_path, current_path, output_path,
               key_override=None, highlight_cols=None,
               fill_color="FFFF00", bold=False, underline=None,
               font_color=None, confirm=None, log=print):
    """Fuehrt den Index-Update aus.

    key_override:   Liste von Header-Namen als Anker (None = Auto-Erkennung)
    highlight_cols: Spaltenangabe fuer die Markierung neuer Zeilen ("B-D",
                    "2-4", None/"all" = alle Spalten)
    fill_color:     Hex-Farbe der Fuellung neuer Zeilen (None = keine Fuellung)
    underline:      None | "single" | "double"
    font_color:     Hex-Schriftfarbe neuer Zeilen (None = unveraendert)
    confirm:        optionaler Callback(liste_von_titeln) -> bool; bei False
                    wird abgebrochen und nichts gespeichert
    Rueckgabe: dict mit existing, new (Liste (zeile, titel)), key, key_reason,
    output -- oder None, wenn confirm abgelehnt hat.
    """
    in_wb, in_ws, in_map, in_rows = read_sheet(input_path)
    cur_wb, cur_ws, cur_map, cur_rows = read_sheet(current_path)

    if key_override:
        key_headers = [norm(h) for h in key_override]
        missing = [h for h in key_headers if h not in in_map or h not in cur_map]
        if missing:
            raise ValueError(f"Anker-Spalte(n) nicht in beiden Dateien vorhanden: {missing}")
        key_reason = "manuell gewaehlt"
        if not (key_is_unique(in_ws, in_map, in_rows, key_headers)
                and key_is_unique(cur_ws, cur_map, cur_rows, key_headers)):
            key_reason += " (WARNUNG: nicht eindeutig)"
    else:
        key_headers, key_reason = choose_key(in_map, in_ws, in_rows,
                                             cur_map, cur_ws, cur_rows)
    log(f"Anker gewaehlt: {key_headers}  ({key_reason})")

    # Vorhandene Schluessel im Current
    cur_keys = {row_key(cur_ws, cur_map, r, key_headers) for r in cur_rows}

    # Spalten-Mapping Input -> Current (per Header-Name)
    col_map = {in_map[h]: cur_map[h] for h in in_map if h in cur_map}

    # Vorlagen fuer Formatierung neuer Zeilen
    col_templates = build_col_templates(cur_ws, cur_map, cur_rows)

    # Ausgabe-Liste aufbauen: bestehende Current-Zeilen als "captured" dicts
    output = [{"type": "existing", "data": capture_row(cur_ws, r)} for r in cur_rows]
    # Zuordnung Current-Key -> Position in output
    key_to_pos = {}
    for i, r in enumerate(cur_rows):
        key_to_pos[row_key(cur_ws, cur_map, r, key_headers)] = i

    # Input in Reihenfolge durchgehen; neue Zeilen hinter ihren Anker einsetzen.
    # da wir waehrend des Einfuegens Positionen verschieben, arbeiten wir mit
    # Marker-Objekten statt Indizes
    output_markers = list(output)  # gleiche Objekte, Reihenfolge = aktuelle Ausgabe

    def find_anchor_pos(anchor_key):
        if anchor_key is None:
            return -1
        for idx, item in enumerate(output_markers):
            if item.get("_key") == anchor_key:
                return idx
        return -1

    # Keys an bestehende Marker haengen
    for item, r in zip(output_markers, cur_rows):
        item["_key"] = row_key(cur_ws, cur_map, r, key_headers)

    last_anchor_key = None
    for r in in_rows:
        k = row_key(in_ws, in_map, r, key_headers)
        if k in cur_keys:
            last_anchor_key = k  # dient als Anker fuer nachfolgende neue Zeilen
        else:
            # neue Zeile: Werte aus Input in Current-Spalten mappen
            new_data = {}
            for c in range(1, cur_ws.max_column + 1):
                new_data[c] = (None, copy(col_templates[c]))
            for in_c, cur_c in col_map.items():
                val = in_ws.cell(r, in_c).value
                new_data[cur_c] = (val, copy(col_templates[cur_c]))
            title = in_ws.cell(r, in_map.get("title", 1)).value
            marker = {"type": "new", "data": new_data, "_key": k, "_title": title}
            pos = find_anchor_pos(last_anchor_key)
            output_markers.insert(pos + 1, marker)
            last_anchor_key = k  # weitere neue Zeilen dahinter, in Reihenfolge

    # Vorschau der neuen Zeilen -> optionale Bestaetigung vor dem Schreiben
    pending_new = [item.get("_title") for item in output_markers
                   if item["type"] == "new"]
    if confirm is not None and not confirm(pending_new):
        log("Abgebrochen -- nichts gespeichert.")
        return None

    # Zielarbeitsmappe = Kopie des Current (behaelt Spaltenbreiten, Blattname etc.)
    out_wb = openpyxl.load_workbook(current_path)
    out_ws = out_wb.active

    # Datenbereich leeren (ab Zeile 2)
    if out_ws.max_row >= 2:
        out_ws.delete_rows(2, out_ws.max_row - 1)

    mark_cols = parse_highlight_cols(highlight_cols, out_ws.max_column)
    fill = None
    if fill_color:
        rgb = fill_color.replace("#", "").upper()[-6:]
        fill = PatternFill(fill_type="solid", fgColor="FF" + rgb)

    # Zeilen neu schreiben; Zeilennummer neuer Zeilen (in der Zieldatei) merken
    new_rows = []  # (zeilennummer_in_ausgabe, title)
    for i, item in enumerate(output_markers):
        row_idx = 2 + i
        is_new = item["type"] == "new"
        if is_new:
            new_rows.append((row_idx, item.get("_title")))
        for c in range(1, out_ws.max_column + 1):
            value, style = item["data"][c]
            cell = out_ws.cell(row_idx, c)
            cell.value = value
            cell._style = copy(style)
            if is_new and c in mark_cols:
                if fill is not None:
                    cell.fill = fill
                if bold or underline or font_color:
                    f = copy(cell.font)
                    if bold:
                        f.bold = True
                    if underline:
                        f.underline = underline
                    if font_color:
                        f.color = openpyxl.styles.Color(
                            rgb="FF" + font_color.replace("#", "").upper()[-6:])
                    cell.font = f

    out_wb.save(output_path)

    log(f"\nBestehende Zeilen: {len(cur_rows)}")
    log(f"Neue Zeilen hinzugefuegt: {len(new_rows)}")
    for row_idx, title in new_rows:
        log(f"  + Zeile {row_idx}: {title}")
    log(f"\nGespeichert: {output_path}")
    return {"existing": len(cur_rows), "new": new_rows,
            "key": key_headers, "key_reason": key_reason,
            "output": output_path}


def main():
    ap = argparse.ArgumentParser(description="Dataroom-Index-Updater")
    ap.add_argument("--input", default="dataroom-input.xlsx",
                    help="Frischer Datenraum-Export")
    ap.add_argument("--current", default="current-index.xlsx",
                    help="Bestehender angereicherter Index")
    ap.add_argument("--output", default="index-updated.xlsx",
                    help="Zieldatei")
    args = ap.parse_args()
    run_update(args.input, args.current, args.output)


if __name__ == "__main__":
    main()
